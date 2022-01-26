# This will be all the functions that actually determine the geared motor selection


import openpyxl
from Classes import *


def find_motors(inputs):  # To find all the applicable motors
    motors = []
    motor_wb = openpyxl.load_workbook('Datasheets/Motors.xlsx')
    for poles in inputs['poles']:
        motor_ws = motor_wb[poles]

        # To find all the data needed for every motor
        sheet_poles = str(motor_ws.cell(column=1, row=1).value[0])
        for r in range(3, motor_ws.max_row+1):
            if motor_ws.cell(column=1, row=r).value == ('' or None):  # This will skip all the invalid rows
                pass
            else:
                sheet_power = float(motor_ws.cell(column=1, row=r).value)
                sheet_speed = float(motor_ws.cell(column=2, row=r).value)
                sheet_frame = str(motor_ws.cell(column=3, row=r).value)
                sheet_shaft = str(motor_ws.cell(column=4, row=r).value)
                sheet_weight = str(motor_ws.cell(column=5, row=r).value)
                sheet_brand = str(motor_ws.cell(column=6, row=r).value)
                sheet_series = str(motor_ws.cell(column=7, row=r).value)

                # To make each motor object and add it to the list
                motor = Motor(sheet_power, sheet_speed, sheet_frame, sheet_shaft, sheet_weight, sheet_brand, sheet_series, sheet_poles)
                motors.append(motor)
    
    return motors


def find_gearboxes(inputs):  # To find all the applicable gearboxes
    gearboxes = []
    p8_cols = [2, 3, 4, 5]
    p6_cols = [6, 7, 8, 9]
    p4_cols = [10, 11, 12, 13]
    p2_cols = [14, 15, 16, 17]
    
    # To open all the appropriate workbooks
    for series in inputs['series_sizes']:
        if len(inputs['series_sizes'][series]) > 0:
            gearbox_wb = openpyxl.load_workbook(f'Datasheets/{series}_Gearboxes.xlsx')
            # To open all the appropriate worksheets in each book
            for size in inputs['series_sizes'][series]:
                gearbox_ws = gearbox_wb[size]

                # Now to find the data for each individual gearbox
                sheet_serie_size = str(gearbox_ws.cell(row=2, column=1).value)
                sheet_shaft = [str(gearbox_ws.cell(row=1, column=15).value), str(gearbox_ws.cell(row=1, column=16).value)]
                # To add all the data to the above empty lists
                for r in range(4, gearbox_ws.max_row+1):
                    if gearbox_ws.cell(column=1, row=r).value == ('' or None):  # This will skip all the invalid rows
                        pass
                    else:
                        sheet_ratio = str(gearbox_ws.cell(row=r, column=1).value)
                        sheet_p8data = []
                        sheet_p6data = []
                        sheet_p4data = []
                        sheet_p2data = []
                        
                        # To populate each of the data points in this list
                        for c in p8_cols:
                            sheet_p8data.append(float(gearbox_ws.cell(row=r, column=c).value))
                        for c in p6_cols:
                            sheet_p6data.append(float(gearbox_ws.cell(row=r, column=c).value))
                        for c in p4_cols:
                            sheet_p4data.append(float(gearbox_ws.cell(row=r, column=c).value))
                        for c in p2_cols:
                            sheet_p2data.append(float(gearbox_ws.cell(row=r, column=c).value))

                        # Now to make each gearbox object
                        gearbox = Gearbox(sheet_serie_size, sheet_ratio, sheet_shaft, sheet_p8data, sheet_p6data, sheet_p4data, sheet_p2data)
                        gearboxes.append(gearbox)

    return gearboxes


def find_gearmotors(inputs):
    motors = find_motors(inputs)
    gearboxes = find_gearboxes(inputs)
    gearmotors = []

    for motor in motors:
        for gearbox in gearboxes:
            gearmotor = GearedMotor(gearbox, motor)
            gearmotors.append(gearmotor)
    
    return gearmotors


def make_recommendations(inputs):
    all_gearmotors = find_gearmotors(inputs)
    applicable_gearmotors = []

    # To turn all the inputs into numbers easy to work with
    spd_lower = inputs['spd'] * (1 - inputs['spd_tol']/100)
    spd_higher = inputs['spd'] * (1 + inputs['spd_tol']/100)
    trq_lower = inputs['trq'] * (1 - inputs['trq_tol']/100)
    trq_higher = inputs['trq'] * (1 + inputs['trq_tol']/100)
    safety_low = inputs['safety_low']
    safety_high = inputs['safety_high']

    for gm in all_gearmotors:
        if spd_lower <= gm.gm_spd <= spd_higher:
            if trq_lower <= gm.gm_trq <= trq_higher:
                if safety_low <= gm.gm_safety <= safety_high:
                    applicable_gearmotors.append(gm)

    return applicable_gearmotors
