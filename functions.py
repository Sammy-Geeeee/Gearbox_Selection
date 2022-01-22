# This will be all the functions that actually determine the geared motor selection


import openpyxl
from classes import *


def interpolation(out_low, out_high, in_low, in_high, in_final):  # Function to interpolate data from the sheets
    out_final = out_low + ((out_high-out_low)/(in_high-in_low))*(in_final-in_low)
    return out_final


def recommend_motor(inputs):  # To recommend a geared motor based on provided inputs
    motor_wb = openpyxl.load_workbook('Motors.xlsx')
    motors = []
    for poles in inputs['poles']:
        motor_ws = motor_wb[poles]

        motors.append(motor_ws.cell(column=2, row=3).value)






    return motors






























# def gearedMotorGiven_motor(inputs):  # To return the motor object, based on the inputs of motor power given
#     # To find and open the correct worksheet
#     motor_wb = openpyxl.load_workbook('Motors.xlsx')
#     if inputs['poles'] == '2':
#         motor_ws = motor_wb['2 Pole']
#     elif inputs['poles'] == '4':
#         motor_ws = motor_wb['4 Pole']
#     elif inputs['poles'] == '6':
#         motor_ws = motor_wb['6 Pole']
#     elif inputs['poles'] == '8':
#         motor_ws = motor_wb['8 Pole']
#     else:
#         print('The worksheet is invalid - Choose appropriate motor poles.')  # TODO - Find a way to keep repeating this until the proper poles are given

#     # To find the correct motor, and all its data
#     for r in range(3, motor_ws.max_row+1):
#         if float(inputs['power']) == float(motor_ws.cell(column=1, row=r).value):
#             sheet_power_cell = motor_ws.cell(column=1, row=r)
#             break
#         else:
#             # print(f'Input power and worksheet power do not line up.')  # TODO - Find a way to keep repeating until a proper motor power is given
#             pass
    
#     sheet_power = float(sheet_power_cell.value)
#     sheet_speed = float(motor_ws.cell(column=2, row=r).value)
#     sheet_frame = str(motor_ws.cell(column=3, row=r).value)
#     sheet_shaft = int(motor_ws.cell(column=4, row=r).value)
#     sheet_weight = float(motor_ws.cell(column=5, row=r).value)
#     sheet_brand = str(motor_ws.cell(column=6, row=r).value)
#     sheet_series = str(motor_ws.cell(column=7, row=r).value)
#     poles = inputs['poles']

#     motor = Motor(sheet_power, sheet_speed, sheet_frame, sheet_shaft, sheet_weight, sheet_brand, sheet_series, poles)
#     return motor


# def gearedMotorGiven_gearboxes(inputs):
#     gearboxes = []

#     # To open the correct workbooks and worksheets
#     series_sizes = inputs['series sizes']
#     for series in series_sizes:
#         if len(series_sizes[series]) > 0:
#             gearbox_wb = openpyxl.load_workbook(f'{series}_Gearboxes.xlsx')
#             for size in series_sizes[series]:
#                 gearbox_ws = gearbox_wb[size]
























# def motorPowerGiven_gearboxes(inputs, motor):  # To determine appropriate gearboxes given the inputs, and a defined motor power
#     gearboxes = []  # To store all our gearbox objects in
    
#     
    
#     # To find the correct columns based on motor speed
#                 motor_speed = motor.speed
#                 if 500 > motor_speed or motor_speed > 2800:
#                     print(f'{motor_speed}    Invalid motor speed')
#                 elif 500 <= motor_speed <= 900:
#                     motor_low = 500
#                     motor_high = 900
#                     cols_higher = [6, 7, 8, 9]
#                     cols_lower = [2, 3, 4, 5]
#                 elif 900 < motor_speed <= 1400:
#                     motor_low = 900
#                     motor_high = 1400
#                     cols_higher = [10, 11, 12, 13]
#                     cols_lower = [6, 7, 8, 9]
#                 elif 1400 < motor_speed <= 2800:
#                     motor_low = 1400
#                     motor_high = 2800
#                     cols_higher = [14, 15, 16, 17]
#                     cols_lower = [10, 11, 12, 13]
                
#                 # To create the gearbox objects we need
#                 series = gearbox_ws.cell(column=1, row=2).value
#                 for r in range(4, gearbox_ws.max_row+1):
                    
#                     # To store the values in each off the appropriate cells
#                     spd_low = gearbox_ws.cell(column=cols_lower[0], row=r).value
#                     trq_low = gearbox_ws.cell(column=cols_lower[1], row=r).value
#                     pwr_low = gearbox_ws.cell(column=cols_lower[2], row=r).value
#                     eff_low = gearbox_ws.cell(column=cols_lower[3], row=r).value
#                     spd_high = gearbox_ws.cell(column=cols_higher[0], row=r).value
#                     trq_high = gearbox_ws.cell(column=cols_higher[1], row=r).value
#                     pwr_high = gearbox_ws.cell(column=cols_higher[2], row=r).value
#                     eff_high = gearbox_ws.cell(column=cols_higher[3], row=r).value

#                     # To find the actual values, after interpolation, and some other values as well
#                     out_spd = interpolation(spd_low, spd_high, motor_low, motor_high, motor_speed)
#                     out_trq = interpolation(trq_low, trq_high, motor_low, motor_high, motor_speed)
#                     out_pwr = interpolation(pwr_low, pwr_high, motor_low, motor_high, motor_speed)
#                     out_eff = interpolation(eff_low, eff_high, motor_low, motor_high, motor_speed)
                    
#                     ratio = gearbox_ws.cell(column=1, row=r).value
#                     gearbox = Gearbox(series, ratio)
#                     gearboxes.append(gearbox)

#     return gearboxes


def gearedMotorGiven_ratings(inputs, motor, gearbox):
    geared_motor = GearedMotor(gearbox, motor)
    geared_motor.printData()
    print()
    
    
    
    
    
    
    # print()
    # print(inputs)
    # print()
    # print(motor)
    # print()
    # print(gearboxes)
    # print()








