# This is just to check that the data is valid


import openpyxl


def checkGearboxData(books):
    # To open up each gearbox sheet
    for book in books:
        wb = openpyxl.load_workbook(f'Datasheets/{book}_Gearboxes.xlsx')
        print(f'\nChecking {book}_Gearboxes.xlsx data...')

        # To iterate through each tab in the sheet
        for ws in wb.worksheets:
            # To go through each row and column containing data in the sheets
            for c in range(2, 17+1):
                for r in range(4, ws.max_row+1):
                    
                    try:
                        cell = ws.cell(column=c, row=r)
                        value = cell.value
                        float(value)  # This is what we're trying to fail
                        # Anything that cannot be converted to a float will be an incorrectly formatted bit of data
                    except Exception as e:
                        print(f'Incorrect Data - {e}')
                        print(f'{ws} - {cell.coordinate} - {value}')
                        print()
        print(f'{book} completed checking.')


def checkMotorData(book):
    wb = openpyxl.load_workbook(f'Datasheets/{book}.xlsx')
    print(f'\n Checking {book}.xlsx data...')

    # To iterate through each tab in the sheet
    for ws in wb.worksheets:
        for r in range(3, ws.max_row+1):
            # To go through all the numeric value columns
            for c in [1, 2, 4, 5]:
                try:
                    cell = ws.cell(column=c, row=r)
                    value = cell.value
                    float(value)
                except Exception as e:
                    print(f'Incorrect Data - {e}')
                    print(f'{ws} - {cell.coordinate} - {value}')
                    print()
            
            # To go through all the string value columns
            for c in [1, 2, 4, 5]:
                try:
                    cell = ws.cell(column=c, row=r)
                    value = cell.value
                    str(value)
                except Exception as e:
                    print(f'Incorrect Data - {e}')
                    print(f'{ws} - {cell.coordinate} - {value}')
                    print()
    print('Motors completed checking.')


gearbox_books = ['VF_W', 'A', 'C', 'F']
motor_book = 'Motors'
checkGearboxData(gearbox_books)
checkMotorData(motor_book)
