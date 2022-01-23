# This will be all the functions that actually determine the geared motor selection


import openpyxl
from Classes import *


def find_motors(inputs):  # To find all the applicable motors
    motors = []
    motor_wb = openpyxl.load_workbook('Motors.xlsx')
    for poles in inputs['poles']:
        motor_ws = motor_wb[poles]

        # To find all the data needed for every motor
        sheet_poles = motor_ws.cell(column=1, row=1).value[0]
        for r in range(3, motor_ws.max_row+1):
            sheet_power = motor_ws.cell(column=1, row=r).value
            sheet_speed = motor_ws.cell(column=2, row=r).value
            sheet_frame = motor_ws.cell(column=3, row=r).value
            sheet_shaft = motor_ws.cell(column=4, row=r).value
            sheet_weight = motor_ws.cell(column=5, row=r).value
            sheet_brand = motor_ws.cell(column=6, row=r).value
            sheet_series = motor_ws.cell(column=7, row=r).value

            # To make each motor object and add it to the list
            motor = Motor(sheet_power, sheet_speed, sheet_frame, sheet_shaft, sheet_weight, sheet_brand, sheet_series, sheet_poles)
            motors.append(motor)
    
    return motors


def find_gearboxes(inputs):  # To find all the applicable gearboxes
    gearboxes = []
    p8_cols = [2, 3, 4, 5]
    p6_cols = [6, 7, 8, 9]
    p4_cols = [10, 11, 12, 13]
    p2_cols = [14, 15, 16, 17]
    # To open all the appropriate workbooks
    for series in inputs['series_sizes']:
        if len(inputs['series_sizes'][series]) > 0:
            gearbox_wb = openpyxl.load_workbook(f'{series}_Gearboxes.xlsx')
            # To open all the appropriate worksheets in each book
            for size in inputs['series_sizes'][series]:
                gearbox_ws = gearbox_wb[size]

                # Now to find the data for each individual gearbox
                sheet_serie_size = gearbox_ws.cell(row=2, column=1).value
                # To add all the data to the above empty lists
                for r in range(4, gearbox_ws.max_row+1):
                    sheet_ratio = gearbox_ws.cell(row=r, column=1).value
                    sheet_p8data = []
                    sheet_p6data = []
                    sheet_p4data = []
                    sheet_p2data = []
                    for c in p8_cols:
                        sheet_p8data.append(gearbox_ws.cell(row=r, column=c).value)
                    for c in p6_cols:
                        sheet_p6data.append(gearbox_ws.cell(row=r, column=c).value)
                    for c in p4_cols:
                        sheet_p4data.append(gearbox_ws.cell(row=r, column=c).value)
                    for c in p2_cols:
                        sheet_p2data.append(gearbox_ws.cell(row=r, column=c).value)

                    # Now to make each gearbox object
                    gearbox = Gearbox(sheet_serie_size, sheet_ratio, sheet_p8data, sheet_p6data, sheet_p4data, sheet_p2data)
                    gearboxes.append(gearbox)

    return gearboxes


def find_gearmotors(inputs):
    motors = find_motors(inputs)
    gearboxes = find_gearboxes(inputs)
    gearmotors = []

    for motor in motors:
        for gearbox in gearboxes:
            gearmotor = GearedMotor(gearbox, motor)
            gearmotors.append(gearmotor)
    
    return gearmotors
